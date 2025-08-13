import pandas as pd
from openpyxl import load_workbook
from .core import normalize_text
from .exceptions import ExcelScannerError
from typing import Optional, Union, List, Tuple
from pathlib import Path


class ExcelScanner:
    """A scanner for Excel files that supports both Pandas (fast bulk operations)
    and OpenPyXL (formula/formatting access)."""

    def __init__(self, filepath: Union[str, Path]) -> None:  # <-- Accepts both str and Path
        """Initialize the Excel scanner with a file path.

        Args:
            filepath: Path to the Excel file (.xlsx, .xls).

        Attributes:
            self.df (pd.DataFrame): Loaded DataFrame (Pandas mode).
            self.wb (Workbook): OpenPyXL workbook object.
            self.ws (Worksheet): Active worksheet (OpenPyXL mode).
        """
        self.filepath = str(filepath)  # type str will backward compatibility with older python path library
        self.df = None  # Private with type hint
        self.wb = None
        self.ws = None

    def load_with_pandas(self) -> pd.DataFrame:
        """Load the Excel file into a Pandas DataFrame.

        Uses header=None to treat all rows as data (no header assumption).
        Caches the DataFrame for subsequent operations.

        Returns:
            pd.DataFrame: The loaded DataFrame.

        Raises:
            ExcelScannerError: If file loading fails.

        Example:
            scanner = ExcelScanner("data.xlsx")
            df = scanner.load_with_pandas()  # Returns DataFrame
        """
        try:
            if self.df is None:
                self.df = pd.read_excel(self.filepath, header=None)
            return self.df
        except Exception as e:
            raise ExcelScannerError(f"Failed to load with pandas: {e}")

    def load_with_openpyxl(self):
        """Load the Excel file using OpenPyXL.

        Enables access to formulas and formatting. Uses data_only=True
        to get calculated values instead of formulas.

        Returns:
            Worksheet: The active worksheet.

        Raises:
            ExcelScannerError: If file loading fails.

        Example:
            ws = scanner.load_with_openpyxl()  # Returns Worksheet
        """
        self.wb = None
        try:
            self.wb = load_workbook(self.filepath, data_only=True)
            self.ws = self.wb.active
            return self.ws
        except Exception as e:
            raise ExcelScannerError(f"Failed to load with openpyxl: {e}")

    def close(self) -> None:
        """Clean up resources, particularly OpenPyXL workbook handles.

        Important to prevent file lock issues on Windows.

        Example:
            scanner.close()  # Release file handles
        """
        if hasattr(self, 'wb') and self.wb:
            self.wb.close()

    def get_cell_content(
            self,
            row: Optional[int] = None,  # Fully optional
            col: Optional[int] = None,  # Fully optional
            get_formula: bool = False,
            row_offset: Optional[int] = None,
            col_offset: Optional[int] = None,
            return_native_type: bool = False,
            use_pandas: bool = True,
            debug: bool = False
    ) -> Union[str, float, int, None]:
        """Get cell content with fully optional positioning and offsets.

        Args:
            row: Base row (1-based). None=first row.
            col: Base column (1-based). None=first column.
            get_formula: Return formula if True (forces OpenPyXL usage).
            row_offset: Offset from base row (None=0).
            col_offset: Offset from base column (None=0).
            return_native_type: If True, returns original data type (str/float/int/None).
                               If False, returns string representation.
            use_pandas: If True, uses Pandas (faster). If False, uses OpenPyXL.
                       Ignored when get_formula=True.
            debug: Print calculation steps.

        Returns:
            Union[str, float, int, None]: Cell content based on return_native_type setting.
            - return_native_type=False: Always returns str, "" for empty/invalid cells
            - return_native_type=True: Returns native type (str/float/int) or None for empty

        Raises:
            ValueError: For negative values or invalid positions.
        """
        # Set defaults (1-based)
        base_row = 1 if row is None else row
        base_col = 1 if col is None else col

        # Convert None offsets to 0
        row_offset = 0 if row_offset is None else row_offset
        col_offset = 0 if col_offset is None else col_offset

        # Validation
        if base_row <= 0 or base_col <= 0:
            raise ValueError(f"Base position must be >=1 (got row={base_row}, col={base_col})")
        if row_offset < 0 or col_offset < 0:
            raise ValueError(f"Offsets must be >=0 (got row_offset={row_offset}, col_offset={col_offset})")

        try:
            target_row = base_row + row_offset
            target_col = base_col + col_offset

            if debug:
                print(f"[DEBUG] Base: ({base_row}, {base_col}) | "
                      f"Offsets: (+{row_offset}, +{col_offset}) -> "
                      f"Target: ({target_row}, {target_col})")

            # Determine which method to use
            use_openpyxl = get_formula or not use_pandas

            if use_openpyxl:
                # Use OpenPyXL for formulas or when explicitly requested
                if not hasattr(self, 'ws') or self.ws is None:
                    self.load_with_openpyxl()

                # OpenPyXL bounds checking
                max_row = self.ws.max_row
                max_col = self.ws.max_column

                if not (1 <= target_row <= max_row and 1 <= target_col <= max_col):
                    if debug:
                        print(f"[DEBUG] Out of bounds (OpenPyXL max: {max_row}x{max_col})")
                    return None if return_native_type else ""

                cell = self.ws.cell(row=target_row, column=target_col)
                value = cell.value

            else:
                # Use Pandas for better performance
                if self.df is None:
                    self.load_with_pandas()

                # Pandas bounds checking
                if not (1 <= target_row <= self.df.shape[0] and
                        1 <= target_col <= self.df.shape[1]):
                    if debug:
                        print(f"[DEBUG] Out of bounds (Pandas max: {self.df.shape})")
                    return None if return_native_type else ""

                value = self.df.iat[target_row - 1, target_col - 1]

            # Handle return type
            if return_native_type:
                # Return native type (None for NaN/empty)
                if pd.isna(value) or value is None:
                    return None
                return value
            else:
                # Return string representation ("" for empty)
                if pd.isna(value) or value is None:
                    return ""
                return str(value)

        except Exception as e:
            if debug:
                import traceback
                print(f"[DEBUG] Error:\n{traceback.format_exc()}")
            return None if return_native_type else ""

    # Optional: Simple wrapper for backward compatibility
    def get_cell_info(self, row: int, col: int, use_pandas: bool = True) -> Union[str, float, int, None]:
        """Get cell content by row and column (1-based indexing) - DEPRECATED.

        Use get_cell_content() with return_native_type=True instead.

        Args:
            row: Row number (1-based).
            col: Column number (1-based).
            use_pandas: If True, uses Pandas (faster). If False, uses OpenPyXL.

        Returns:
            The cell value (type depends on content - str, float, int, None)
        """
        return self.get_cell_content(
            row=row,
            col=col,
            return_native_type=True,
            use_pandas=use_pandas
        )

    def get_keyword_cell(
            self,
            keyword: str,
            exact_match: bool = True,
            end_row: Optional[int] = None,
            end_col: Optional[int] = None
    ) -> List[Tuple[int, int]]:
        """Find all cells containing the keyword (exact or partial match).

        Performs case-insensitive search after normalizing both the
        keyword and cell contents. Returns 1-based coordinates.

        Args:
            keyword: The search term to find.
            exact_match: If True, looks for exact matches only. If False,
                looks for partial matches (cell contains keyword).
            end_row: Last row to search (1-based). If None, searches all rows.
            end_col: Last column to search (1-based). If None, searches all columns.

        Returns:
            List of (row, col) tuples where keyword was found.

        Example:
            scanner.get_keyword_cell("Total", end_row=10, end_col=5)
            [(1, 3), (5, 2)]  # Found at row 1 col 3 and row 5 col 2
        """
        if self.df is None:
            self.load_with_pandas()

        # Convert 1-based to 0-based and handle None
        row_slice = slice(None, end_row) if end_row is None else slice(None, end_row - 1)
        col_slice = slice(None, end_col) if end_col is None else slice(None, end_col - 1)

        # Get the subset of dataframe to search
        search_area = self.df.iloc[row_slice, col_slice]

        norm_keyword = normalize_text(str(keyword))

        if exact_match:
            matches = search_area.map(lambda x: normalize_text(str(x)) == norm_keyword)
        else:
            matches = search_area.map(lambda x: norm_keyword in normalize_text(str(x)))

        rows, cols = matches.to_numpy().nonzero()
        return [(int(row + 1), int(col + 1)) for row, col in zip(rows, cols)]

    def find_consensus_row(
            self,
            keywords: set[str],
            exact_match: bool = True,
            end_row: Optional[int] = None,
            end_col: Optional[int] = None
    ) -> Optional[int]:
        """Find the single row where ALL keywords appear.

        Args:
            keywords: Set of keywords. For accurate results, provide 2+ keywords.
            exact_match: If True, requires exact matches.
            end_row: Limit row search range.
            end_col: Limit col search range.

        Returns:
            int: Common row number if all keywords share exactly one row.
            None: If any keyword has no matches.

        Raises:
            ValueError: With specific guidance when:
                - Single keyword matches multiple rows (suggests adding more keywords)
                - Keywords don't share any row
                - Keywords share multiple rows
        """
        if not keywords:
            return None

        # Step 1: Get all rows for each keyword
        keyword_to_rows = {}
        for keyword in keywords:
            positions = self.get_keyword_cell(keyword, exact_match, end_row, end_col)
            if not positions:
                return None
            keyword_to_rows[keyword] = {pos[0] for pos in positions}

        # Step 2: Special case - single keyword with multiple matches
        if len(keywords) == 1 and len(next(iter(keyword_to_rows.values()))) > 1:
            keyword = next(iter(keywords))
            matched_rows = keyword_to_rows[keyword]
            raise ValueError(
                f"Keyword '{keyword}' appears in multiple rows: {matched_rows}. "
                f"Please provide 1-2 additional keywords to identify the correct row.\n"
                f"Example: scanner.find_consensus_row({{{keyword!r}, 'AdditionalKeyword1', 'AdditionalKeyword2'}})"
            )

        # Step 3: Multi-keyword validation
        common_rows = set.intersection(*keyword_to_rows.values())

        if not common_rows:
            all_matches = {k: v for k, v in keyword_to_rows.items()}
            raise ValueError(
                f"No common row found. Keywords appear in different rows: {all_matches}\n"
                f"Suggestions:\n"
                f"1. Verify keyword spelling\n"
                f"2. Add more specific keywords\n"
                f"3. Adjust search range (end_row/end_col)"
            )
        elif len(common_rows) > 1:
            raise ValueError(
                f"Keywords appear in multiple shared rows: {common_rows}\n"
                f"Try narrowing the search with:\n"
                f"1. More specific keywords\n"
                f"2. Smaller end_row/end_col range"
            )

        return common_rows.pop()

    def find_consensus_col(
            self,
            keywords: set[str],
            exact_match: bool = True,
            end_row: Optional[int] = None,
            end_col: Optional[int] = None
    ) -> Optional[int]:
        """Find the single column where ALL keywords appear.

        Args:
            keywords: Set of keywords. For accurate results, provide 2+ keywords.
            exact_match: If True, requires exact matches.
            end_row: Limit row search range.
            end_col: Limit col search range.

        Returns:
            int: Common column number if all keywords share exactly one column.
            None: If any keyword has no matches.

        Raises:
            ValueError: With specific guidance when:
                - Single keyword matches multiple columns (suggests adding more keywords)
                - Keywords don't share any column
                - Keywords share multiple columns
        """
        if not keywords:
            return None

        # Step 1: Get all columns for each keyword
        keyword_to_cols = {}
        for keyword in keywords:
            positions = self.get_keyword_cell(keyword, exact_match, end_row, end_col)
            if not positions:
                return None
            keyword_to_cols[keyword] = {pos[1] for pos in positions}

        # Step 2: Special case - single keyword with multiple matches
        if len(keywords) == 1 and len(next(iter(keyword_to_cols.values()))) > 1:
            keyword = next(iter(keywords))
            matched_cols = keyword_to_cols[keyword]
            raise ValueError(
                f"Keyword '{keyword}' appears in multiple columns: {matched_cols}. "
                f"Please provide 1-2 additional keywords to identify the correct column.\n"
                f"Example: scanner.find_consensus_col({{{keyword!r}, 'AdditionalKeyword1', 'AdditionalKeyword2'}})"
            )

        # Step 3: Multi-keyword validation
        common_cols = set.intersection(*keyword_to_cols.values())

        if not common_cols:
            all_matches = {k: v for k, v in keyword_to_cols.items()}
            raise ValueError(
                f"No common column found. Keywords appear in different columns: {all_matches}\n"
                f"Suggestions:\n"
                f"1. Verify keyword spelling\n"
                f"2. Add more specific keywords\n"
                f"3. Adjust search range (end_row/end_col)"
            )
        elif len(common_cols) > 1:
            raise ValueError(
                f"Keywords appear in multiple shared columns: {common_cols}\n"
                f"Try narrowing the search with:\n"
                f"1. More specific keywords\n"
                f"2. Smaller end_row/end_col range"
            )

        return common_cols.pop()

    def get_slice_content(
            self,
            start_slice_row: Optional[int] = None,
            end_slice_row: Optional[int] = None,
            start_slice_col: Optional[int] = None,
            end_slice_col: Optional[int] = None
    ) -> pd.DataFrame:
        """Extract a slice of the Excel content as a pandas DataFrame.

        Args:
            start_slice_row: First row to include (1-based). None = start from row 1.
            end_slice_row: Last row to include (1-based). None = end at last data row.
            start_slice_col: First column to include (1-based). None = start from column 1.
            end_slice_col: Last column to include (1-based). None = end at last data column.

        Returns:
            pd.DataFrame: Sliced data preserving original pandas format.

        Example:
            # Get rows 5-10, all columns
            df = scanner.get_slice_content(start_slice_row=5, end_slice_row=10)

            # Get all rows, columns 3-5
            df = scanner.get_slice_content(start_slice_col=3, end_slice_col=5)
        """
        if self.df is None:
            self.load_with_pandas()

        # Convert 1-based to 0-based and handle None values
        start_row = 0 if start_slice_row is None else start_slice_row - 1
        end_row = len(self.df) if end_slice_row is None else end_slice_row
        start_col = 0 if start_slice_col is None else start_slice_col - 1
        end_col = len(self.df.columns) if end_slice_col is None else end_slice_col

        # Perform slicing (iloc is exclusive on end index)
        return self.df.iloc[start_row:end_row, start_col:end_col].copy()

'''
    def get_cell_info(self, row: int, col: int, use_pandas: bool = True) -> Union[str, float, int]:
        """Get cell content by row and column (1-based indexing).

        Args:
            row: Row number (1-based).
            col: Column number (1-based).
            use_pandas: If True, uses Pandas (faster). If False, uses OpenPyXL.

        Returns:
            The cell value (type depends on content - str, float, int, etc.)

        Example:
            scanner.get_cell_info(1, 1)  # Returns content of A1
            "Header"
        """
        if use_pandas:
            if not hasattr(self, 'df'):
                self.load_with_pandas()
            return self.df.iloc[row - 1, col - 1]
        else:
            if not hasattr(self, 'ws'):
                self.load_with_openpyxl()
            return self.ws.cell(row=row, column=col).value

    def get_cell_content(
            self,
            row: Optional[int] = None,  # Fully optional
            col: Optional[int] = None,  # Fully optional
            get_formula: bool = False,
            row_offset: Optional[int] = None,
            col_offset: Optional[int] = None,
            debug: bool = False
    ) -> str:
        """Get cell content with fully optional positioning and offsets.

        Args:
            row: Base row (1-based). None=first row.
            col: Base column (1-based). None=first column.
            get_formula: Return formula if True.
            row_offset: Offset from base row (None=0).
            col_offset: Offset from base column (None=0).
            debug: Print calculation steps.

        Returns:
            str: Cell content or "" if invalid.

        Raises:
            ValueError: For negative values or invalid positions.
        """
        # Set defaults (1-based)
        base_row = 1 if row is None else row
        base_col = 1 if col is None else col

        # Convert None offsets to 0
        row_offset = 0 if row_offset is None else row_offset
        col_offset = 0 if col_offset is None else col_offset

        # Validation
        if base_row <= 0 or base_col <= 0:
            raise ValueError(f"Base position must be >=1 (got row={base_row}, col={base_col})")
        if row_offset < 0 or col_offset < 0:
            raise ValueError(f"Offsets must be >=0 (got row_offset={row_offset}, col_offset={col_offset})")

        try:
            if self.df is None:
                self.load_with_pandas()

            target_row = base_row + row_offset
            target_col = base_col + col_offset

            if debug:
                print(f"[DEBUG] Base: ({base_row}, {base_col}) | "
                      f"Offsets: (+{row_offset}, +{col_offset}) -> "
                      f"Target: ({target_row}, {target_col})")

            # Boundary check
            if not (1 <= target_row <= self.df.shape[0] and
                    1 <= target_col <= self.df.shape[1]):
                if debug:
                    print(f"[DEBUG] Out of bounds (max {self.df.shape})")
                return ""

            # Get content
            if get_formula:
                if not hasattr(self, 'ws'):
                    self.load_with_openpyxl()
                value = self.ws.cell(target_row, target_col).value
            else:
                value = self.df.iat[target_row - 1, target_col - 1]

            return str(value) if pd.notna(value) else ""

        except Exception as e:
            if debug:
                import traceback
                print(f"[DEBUG] Error:\n{traceback.format_exc()}")
            return ""

'''
''' V2
    def find_consensus_row(
            self,
            keywords: set[str],
            exact_match: bool = True,
            end_row: Optional[int] = None,
            end_col: Optional[int] = None
    ) -> Optional[int]:
        """Find the single row where ALL keywords appear (may be in different columns).

        For multiple keywords (2+), ALL must share at least one common row.

        Args:
            keywords: Set of keywords (e.g., {"Product", "Brand"}).
            exact_match: If True, requires exact matches.
            end_row/end_col: Limit search range.

        Returns:
            int: Common row number if all keywords share at least one row.
            None: If any keyword has no matches.

        Raises:
            ValueError: If keywords don't share any row or disagree.
        """
        if not keywords:
            return None

        # Step 1: Get all rows for each keyword
        keyword_to_rows = {}
        for keyword in keywords:
            positions = self.get_keyword_cell(
                keyword=keyword,
                exact_match=exact_match,
                end_row=end_row,
                end_col=end_col
            )
            if not positions:
                return None  # Early exit if any keyword is missing
            keyword_to_rows[keyword] = {pos[0] for pos in positions}

        # Step 2: Find intersection of all rows
        common_rows = set.intersection(*keyword_to_rows.values())

        if not common_rows:
            raise ValueError(
                f"No common row found. Keywords appear in different rows: "
                f"{ {k: v for k, v in keyword_to_rows.items()} }"
            )
        elif len(common_rows) > 1:
            raise ValueError(
                f"Keywords appear in multiple shared rows: {common_rows}. "
                f"Expected exactly one common row."
            )

        return common_rows.pop()

    def find_consensus_col(
            self,
            keywords: set[str],
            exact_match: bool = True,
            end_row: Optional[int] = None,
            end_col: Optional[int] = None
    ) -> Optional[int]:
        """Find the single column where ALL keywords appear (may be in different rows).

        For multiple keywords (2+), ALL must share at least one common column.

        Args:
            keywords: Set of keywords (e.g., {"Total", "Subtotal"}).
            exact_match: If True, requires exact matches.
            end_row/end_col: Limit search range.

        Returns:
            int: Common column number if all keywords share at least one column.
            None: If any keyword has no matches.

        Raises:
            ValueError: If keywords don't share any column or disagree.
        """
        if not keywords:
            return None

        # Step 1: Get all columns for each keyword
        keyword_to_cols = {}
        for keyword in keywords:
            positions = self.get_keyword_cell(
                keyword=keyword,
                exact_match=exact_match,
                end_row=end_row,
                end_col=end_col
            )
            if not positions:
                return None  # Early exit if any keyword is missing
            keyword_to_cols[keyword] = {pos[1] for pos in positions}

        # Step 2: Find intersection of all columns
        common_cols = set.intersection(*keyword_to_cols.values())

        if not common_cols:
            raise ValueError(
                f"No common column found. Keywords appear in different columns: "
                f"{ {k: v for k, v in keyword_to_cols.items()} }"
            )
        elif len(common_cols) > 1:
            raise ValueError(
                f"Keywords appear in multiple shared columns: {common_cols}. "
                f"Expected exactly one common column."
            )

        return common_cols.pop()
'''

'''
    def find_consensus_row(
            self,
            keywords: set[str],
            exact_match: bool = True,
            end_row: Optional[int] = None,
            end_col: Optional[int] = None
    ) -> Optional[int]:
        """Find the single row where ALL keywords appear (may be in different columns).

        Args:
            keywords: Set of keywords (e.g., {"Product", "Brand"}).
            exact_match: If True, requires exact matches.
            end_row/end_col: Limit search range.

        Returns:
            int: Common row number (e.g., 25) if all keywords share it.
            None: If no consensus or no matches.

        Raises:
            ValueError: If keywords appear in different rows.
        """
        if not keywords:
            return None

        # Step 1: Get all (row, col) positions for each keyword
        keyword_rows = set()
        for keyword in keywords:
            positions = self.get_keyword_cell(
                keyword=keyword,
                exact_match=exact_match,
                end_row=end_row,
                end_col=end_col
            )
            if not positions:
                return None  # At least one keyword has no matches
            keyword_rows.update({pos[0] for pos in positions})

        # Step 2: Check if all keywords share a single row
        if len(keyword_rows) == 1:
            return keyword_rows.pop()
        else:
            raise ValueError(f"Keywords disagree on rows: {keyword_rows}")


    def find_consensus_col(
            self,
            keywords: set[str],
            exact_match: bool = True,
            end_row: Optional[int] = None,
            end_col: Optional[int] = None
    ) -> Optional[int]:
        """Find the single column where ALL keywords appear (may be in different rows).

        Args:
            keywords: Set of keywords (e.g., {"Total", "Subtotal"}).
            exact_match: If True, requires exact matches.
            end_row/end_col: Limit search range.

        Returns:
            int: Common column number (e.g., 17) if all keywords share it.
            None: If no consensus or no matches.

        Raises:
            ValueError: If keywords appear in different columns.
        """
        if not keywords:
            return None

        # Step 1: Get all (row, col) positions for each keyword
        keyword_cols = set()
        for keyword in keywords:
            positions = self.get_keyword_cell(
                keyword=keyword,
                exact_match=exact_match,
                end_row=end_row,
                end_col=end_col
            )
            if not positions:
                return None  # At least one keyword has no matches
            keyword_cols.update({pos[1] for pos in positions})

        # Step 2: Check if all keywords share a single column
        if len(keyword_cols) == 1:
            return keyword_cols.pop()
        else:
            raise ValueError(f"Keywords disagree on columns: {keyword_cols}")

'''
